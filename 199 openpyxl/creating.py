# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook() # Criando uma instãncia 
# worksheet: Worksheet = workbook.active # Ativando e jogando no worksheet


sheet_name = 'Minha planilha' # Nome para a planilha
workbook.create_sheet(sheet_name, 0) # Criamos a planilha
worksheet: Worksheet = workbook[sheet_name] # Selecionou a planilha
# print(workbook.sheetnames)

# Remover uma planilha
workbook.remove(workbook['Sheet']) # Sheet sempre é criado ao criar uma planilha
 
# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome') # Linha 1 Coluna 1
worksheet.cell(1, 2, 'Idade') # Linha 1 Coluna 2
worksheet.cell(1, 3, 'Nota') # Linha 1 Coluna 3

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i in range(2, 10):
#     for j in range(1, 4):
#         print('Linha', i, 'Coluna', j)

# Lógica complicada
# for i, student_row in enumerate(students, start=2):
#     # print(i, student_row) # Linha 
#     for j, student_column in enumerate(student_row, start=1):
#         # print(i, j, student_column) # Coluna
#         worksheet.cell(i, j, student_column)

# Lógica fácil
for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH) # Salvar na pasta WORKBOOK_PATH